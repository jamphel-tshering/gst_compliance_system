from django.contrib import admin
from django.shortcuts import render, redirect
from django.urls import reverse
from django.http import HttpResponse, FileResponse
from django.utils import timezone
from datetime import datetime
import csv
import io
from .models import ReportTemplate, GeneratedReport, ReportSchedule, DashboardWidget, AnalyticsData


@admin.register(ReportTemplate)
class ReportTemplateAdmin(admin.ModelAdmin):
    list_display = ['name', 'category', 'report_type', 'is_active', 'is_public', 'created_at']
    list_filter = ['category', 'report_type', 'is_active', 'is_public']
    search_fields = ['name', 'description']
    



@admin.register(GeneratedReport)
class GeneratedReportAdmin(admin.ModelAdmin):
    list_display = ['report_name', 'report_template', 'report_status', 'generated_by', 'generated_at', 'download_report', 'print_report']
    list_filter = ['report_status', 'file_type', 'generated_at']
    search_fields = ['report_name', 'report_template__name']
    readonly_fields = ['generated_at', 'file_size']
    actions = ['generate_sample_csv_report', 'generate_sample_excel_report']
    change_list_template = 'admin/generated_report_change_list.html'
    
    def get_urls(self):
        from django.urls import path
        urls = super().get_urls()
        custom_urls = [
            path('generate/taxpayer-csv/', self.admin_site.admin_view(self.generate_taxpayer_csv), name='generate_taxpayer_csv'),
            path('generate/taxpayer-excel/', self.admin_site.admin_view(self.generate_taxpayer_excel), name='generate_taxpayer_excel'),
            path('generate/returns-csv/', self.admin_site.admin_view(self.generate_returns_csv), name='generate_returns_csv'),
            path('generate/returns-excel/', self.admin_site.admin_view(self.generate_returns_excel), name='generate_returns_excel'),
            path('generate/compliance-csv/', self.admin_site.admin_view(self.generate_compliance_csv), name='generate_compliance_csv'),
            path('generate/compliance-excel/', self.admin_site.admin_view(self.generate_compliance_excel), name='generate_compliance_excel'),
        ]
        return custom_urls + urls
    
    def download_report(self, obj):
        """Download button for report"""
        if obj.output_file:
            return f'<a href="{obj.output_file.url}" class="button" download>Download</a>'
        return 'No file'
    download_report.short_description = 'Download'
    download_report.allow_tags = True
    
    def print_report(self, obj):
        """Print button for report"""
        if obj.output_file:
            return f'<button onclick="window.open(\'{obj.output_file.url}\', \'_blank\'); window.print();" class="button">Print</button>'
        return 'No file'
    print_report.short_description = 'Print'
    print_report.allow_tags = True
    
    def generate_sample_csv_report(self, request, queryset):
        """Generate a sample CSV report"""
        from taxpayers.models import TaxpayerMaster
        
        # Create CSV content
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="taxpayer_report.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['GSTIN', 'Taxpayer Name', 'Business Name', 'Status', 'Dzongkhag', 'Organisation Type'])
        
        taxpayers = TaxpayerMaster.objects.filter(is_primary_license=True)[:100]
        for taxpayer in taxpayers:
            writer.writerow([
                taxpayer.gstin,
                taxpayer.taxpayer_name,
                taxpayer.business_name,
                taxpayer.status,
                taxpayer.dzongkhag,
                taxpayer.organisation_type
            ])
        
        return response
    
    generate_sample_csv_report.short_description = 'Generate Sample CSV Report'
    
    def generate_sample_excel_report(self, request, queryset):
        """Generate a sample Excel report"""
        from taxpayers.models import TaxpayerMaster
        from django.http import HttpResponse
        import openpyxl
        from openpyxl.styles import Font
        from django.core.files.base import ContentFile
        from django.utils.text import slugify
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Taxpayer Report'
        
        # Add headers
        headers = ['GSTIN', 'Taxpayer Name', 'Business Name', 'Status', 'Dzongkhag', 'Organisation Type']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        
        # Add data
        taxpayers = TaxpayerMaster.objects.filter(is_primary_license=True)[:100]
        for row, taxpayer in enumerate(taxpayers, 2):
            ws.cell(row=row, column=1, value=taxpayer.gstin)
            ws.cell(row=row, column=2, value=taxpayer.taxpayer_name)
            ws.cell(row=row, column=3, value=taxpayer.business_name)
            ws.cell(row=row, column=4, value=taxpayer.status)
            ws.cell(row=row, column=5, value=taxpayer.dzongkhag)
            ws.cell(row=row, column=6, value=taxpayer.organisation_type)
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create response
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="taxpayer_report.xlsx"'
        
        return response
    
    generate_sample_excel_report.short_description = 'Generate Sample Excel Report'
    
    def generate_taxpayer_csv(self, request):
        """Direct URL method to generate taxpayer CSV report"""
        from taxpayers.models import TaxpayerMaster
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="taxpayer_report.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['GSTIN', 'Taxpayer Name', 'Business Name', 'Status', 'Dzongkhag', 'Organisation Type', 'Registration Date'])
        
        taxpayers = TaxpayerMaster.objects.filter(is_primary_license=True)
        for taxpayer in taxpayers:
            writer.writerow([
                taxpayer.gstin,
                taxpayer.taxpayer_name,
                taxpayer.business_name,
                taxpayer.status,
                taxpayer.dzongkhag,
                taxpayer.organisation_type,
                taxpayer.registration_date.strftime('%d-%m-%Y') if taxpayer.registration_date else ''
            ])
        
        return response
    
    def generate_taxpayer_excel(self, request):
        """Direct URL method to generate taxpayer Excel report"""
        from taxpayers.models import TaxpayerMaster
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Taxpayer Report'
        
        headers = ['GSTIN', 'Taxpayer Name', 'Business Name', 'Status', 'Dzongkhag', 'Organisation Type', 'Registration Date']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        
        taxpayers = TaxpayerMaster.objects.filter(is_primary_license=True)
        for row, taxpayer in enumerate(taxpayers, 2):
            ws.cell(row=row, column=1, value=taxpayer.gstin)
            ws.cell(row=row, column=2, value=taxpayer.taxpayer_name)
            ws.cell(row=row, column=3, value=taxpayer.business_name)
            ws.cell(row=row, column=4, value=taxpayer.status)
            ws.cell(row=row, column=5, value=taxpayer.dzongkhag)
            ws.cell(row=row, column=6, value=taxpayer.organisation_type)
            ws.cell(row=row, column=7, value=taxpayer.registration_date.strftime('%d-%m-%Y') if taxpayer.registration_date else '')
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="taxpayer_report.xlsx"'
        
        return response
    
    def generate_returns_csv(self, request):
        """Direct URL method to generate returns CSV report"""
        from returns.models import GSTReturn
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="gst_returns_report.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Tax Period', 'GSTIN', 'Taxpayer Name', 'Declared Sales', 'Declared Output GST', 'GST Payable/Refundable', 'Filing Status', 'Payment Status'])
        
        returns = GSTReturn.objects.all()
        for ret in returns:
            writer.writerow([
                ret.tax_period,
                ret.gstin,
                ret.taxpayer_name,
                ret.declared_sales,
                ret.declared_output_gst,
                ret.gst_payable_refundable,
                ret.filing_status,
                ret.payment_status
            ])
        
        return response
    
    def generate_returns_excel(self, request):
        """Direct URL method to generate returns Excel report"""
        from returns.models import GSTReturn
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'GST Returns Report'
        
        headers = ['Tax Period', 'GSTIN', 'Taxpayer Name', 'Declared Sales', 'Declared Output GST', 'GST Payable/Refundable', 'Filing Status', 'Payment Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        
        returns = GSTReturn.objects.all()
        for row, ret in enumerate(returns, 2):
            ws.cell(row=row, column=1, value=ret.tax_period)
            ws.cell(row=row, column=2, value=ret.gstin)
            ws.cell(row=row, column=3, value=ret.taxpayer_name)
            ws.cell(row=row, column=4, value=ret.declared_sales)
            ws.cell(row=row, column=5, value=ret.declared_output_gst)
            ws.cell(row=row, column=6, value=ret.gst_payable_refundable)
            ws.cell(row=row, column=7, value=ret.filing_status)
            ws.cell(row=row, column=8, value=ret.payment_status)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="gst_returns_report.xlsx"'
        
        return response
    
    def generate_compliance_csv(self, request):
        """Direct URL method to generate compliance CSV report"""
        from compliance.models import ComplianceMonitoring
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="compliance_report.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Compliance ID', 'Tax Period', 'GSTIN', 'Taxpayer Name', 'Filing Status', 'Filing Delay', 'Payment Status', 'Compliance Status'])
        
        compliance_records = ComplianceMonitoring.objects.all()
        for record in compliance_records:
            writer.writerow([
                record.compliance_id,
                record.tax_period,
                record.gstin,
                record.taxpayer_name,
                record.filing_status,
                record.filing_delay,
                record.payment_status,
                record.compliance_status
            ])
        
        return response
    
    def generate_compliance_excel(self, request):
        """Direct URL method to generate compliance Excel report"""
        from compliance.models import ComplianceMonitoring
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Compliance Report'
        
        headers = ['Compliance ID', 'Tax Period', 'GSTIN', 'Taxpayer Name', 'Filing Status', 'Filing Delay', 'Payment Status', 'Compliance Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        
        compliance_records = ComplianceMonitoring.objects.all()
        for row, record in enumerate(compliance_records, 2):
            ws.cell(row=row, column=1, value=record.compliance_id)
            ws.cell(row=row, column=2, value=record.tax_period)
            ws.cell(row=row, column=3, value=record.gstin)
            ws.cell(row=row, column=4, value=record.taxpayer_name)
            ws.cell(row=row, column=5, value=record.filing_status)
            ws.cell(row=row, column=6, value=record.filing_delay)
            ws.cell(row=row, column=7, value=record.payment_status)
            ws.cell(row=row, column=8, value=record.compliance_status)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="compliance_report.xlsx"'
        
        return response


@admin.register(ReportSchedule)
class ReportScheduleAdmin(admin.ModelAdmin):
    list_display = ['schedule_name', 'report_template', 'frequency', 'next_run', 'last_run', 'is_active']
    list_filter = ['frequency', 'is_active']
    search_fields = ['schedule_name', 'report_template__name']


@admin.register(DashboardWidget)
class DashboardWidgetAdmin(admin.ModelAdmin):
    list_display = ['name', 'widget_type', 'chart_type', 'position_x', 'position_y', 'is_active']
    list_filter = ['widget_type', 'chart_type', 'is_active']
    search_fields = ['name', 'data_source']


@admin.register(AnalyticsData)
class AnalyticsDataAdmin(admin.ModelAdmin):
    list_display = ['data_key', 'updated_at', 'expires_at']
    search_fields = ['data_key']
    readonly_fields = ['data_key', 'data_value', 'created_at', 'updated_at']


# Admin dashboard view for reporting module
def reporting_dashboard(request):
    """Reporting Module Dashboard"""
    context = {
        'title': 'Reporting Module',
        'subtitle': 'Centralized Reporting and Analytics Layer',
        'dashboard_url': '/reports/',
        'report_templates_url': reverse('admin:reporting_reporttemplate_changelist'),
        'generated_reports_url': reverse('admin:reporting_generatedreport_changelist'),
        'report_schedules_url': reverse('admin:reporting_reportschedule_changelist'),
        'dashboard_widgets_url': reverse('admin:reporting_dashboardwidget_changelist'),
    }
    return render(request, 'reporting/admin_dashboard.html', context)