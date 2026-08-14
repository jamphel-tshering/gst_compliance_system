from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.admin.views.decorators import staff_member_required
from django.http import JsonResponse, HttpResponse
from django.db.models import Count, Sum, Avg, Q, F, Case, When, Value, IntegerField
from django.db.models.functions import Coalesce
from django.utils import timezone
from django.contrib import messages
from django.core.paginator import Paginator
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import csv
import io
import time
from .models import ReportTemplate, GeneratedReport, ReportSchedule
from taxpayers.models import TaxpayerMaster
from returns.models import GSTReturn
from compliance.models import ComplianceMonitoring, ComplianceRiskReferral, EnforcementRecovery
from audit_refund.models import AuditCase, AuditAssessment, AuditFinding, RefundRegister
from core.models import User


# Common Reporting Engine
class ReportEngine:
    """Centralized reporting engine for all report types"""
    
    def __init__(self):
        self.tax_periods = self._get_tax_periods()
        self.dzongkhags = self._get_dzongkhags()
        self.sectors = self._get_sectors()
        self.organisation_types = self._get_organisation_types()
        self.risk_levels = ['Low', 'Medium', 'High', 'Critical']
        self.audit_statuses = ['Pending Assignment', 'Assigned', 'In Progress', 'Completed', 'Closed']
        self.refund_statuses = ['submitted', 'under_review', 'processing', 'approved', 'rejected', 'paid', 'closed']
    
    def _get_tax_periods(self):
        """Get available tax periods from GST Returns"""
        periods = GSTReturn.objects.values_list('tax_period', flat=True).distinct().order_by('tax_period')
        return list(periods)
    
    def _get_dzongkhags(self):
        """Get available dzongkhags from taxpayers"""
        dzongkhags = TaxpayerMaster.objects.values_list('dzongkhag', flat=True).distinct().order_by('dzongkhag')
        return [d for d in dzongkhags if d]
    
    def _get_sectors(self):
        """Get available sectors from taxpayers"""
        sectors = TaxpayerMaster.objects.values_list('sector', flat=True).distinct().order_by('sector')
        return [s for s in sectors if s]
    
    def _get_organisation_types(self):
        """Get available organisation types from taxpayers"""
        org_types = TaxpayerMaster.objects.values_list('organisation_type', flat=True).distinct().order_by('organisation_type')
        return [o for o in org_types if o]
    
    def apply_filters(self, queryset, filters):
        """Apply common filters to queryset"""
        if filters.get('from_tax_period'):
            queryset = queryset.filter(tax_period__gte=filters['from_tax_period'])
        if filters.get('to_tax_period'):
            queryset = queryset.filter(tax_period__lte=filters['to_tax_period'])
        if filters.get('gstin'):
            queryset = queryset.filter(gstin__icontains=filters['gstin'])
        if filters.get('taxpayer_name'):
            queryset = queryset.filter(taxpayer_name__icontains=filters['taxpayer_name'])
        if filters.get('dzongkhag'):
            queryset = queryset.filter(dzongkhag=filters['dzongkhag'])
        if filters.get('sector'):
            queryset = queryset.filter(sector=filters['sector'])
        if filters.get('organisation_type'):
            queryset = queryset.filter(organisation_type=filters['organisation_type'])
        if filters.get('risk_level'):
            queryset = queryset.filter(risk_level=filters['risk_level'])
        if filters.get('status'):
            queryset = queryset.filter(status=filters['status'])
        
        return queryset
    
    def format_number(self, value, decimal_places=2):
        """Format numbers for display"""
        if value is None:
            return 0
        return round(float(value), decimal_places)
    
    def calculate_percentage(self, numerator, denominator):
        """Calculate percentage safely"""
        if denominator and denominator != 0:
            return round((numerator / denominator) * 100, 2)
        return 0


# Management Reports
@staff_member_required
def report_landing(request):
    """Professional Report Landing Page"""
    report_categories = [
        {
            'id': 'management',
            'name': 'Management Reports',
            'icon': '📊',
            'description': 'Executive summaries and KPI dashboards',
            'reports': [
                {'id': 'executive_summary', 'name': 'Executive GST Summary', 'description': 'High-level GST overview'},
            ]
        },
        {
            'id': 'taxpayer',
            'name': 'Taxpayer Reports',
            'icon': '👥',
            'description': 'Taxpayer registration and master data',
            'reports': [
                {'id': 'taxpayer_master', 'name': 'Taxpayer Master Report', 'description': 'Complete taxpayer list'},
                {'id': 'new_registrations', 'name': 'New Registration Report', 'description': 'Recent taxpayer registrations'},
                {'id': 'deregistrations', 'name': 'Deregistration Report', 'description': 'Taxpayer deregistrations'},
                {'id': 'taxpayer_by_dzongkhag', 'name': 'Taxpayer by Dzongkhag', 'description': 'Geographic distribution'},
                {'id': 'taxpayer_by_sector', 'name': 'Taxpayer by Sector', 'description': 'Sector-wise distribution'},
                {'id': 'taxpayer_by_org_type', 'name': 'Taxpayer by Organisation Type', 'description': 'Organisation type distribution'},
                {'id': 'taxpayer_by_frequency', 'name': 'Taxpayer by Filing Frequency', 'description': 'Filing frequency distribution'},
            ]
        },
        {
            'id': 'returns',
            'name': 'GST Return & Revenue Reports',
            'icon': '💰',
            'description': 'Return filing and revenue analysis',
            'reports': [
                {'id': 'return_filing_summary', 'name': 'Return Filing Summary', 'description': 'Filing compliance overview'},
                {'id': 'return_filing_detail', 'name': 'Return Filing Detail', 'description': 'Detailed filing information'},
                {'id': 'gst_declaration_summary', 'name': 'GST Declaration Summary', 'description': 'Revenue declaration summary'},
                {'id': 'payment_compliance', 'name': 'Payment Compliance Report', 'description': 'Payment status analysis'},
                {'id': 'negative_returns', 'name': 'Negative Return Report', 'description': 'Negative return analysis'},
                {'id': 'return_amendments', 'name': 'Return Amendment Report', 'description': 'Amendment tracking'},
            ]
        },
        {
            'id': 'compliance',
            'name': 'Compliance Reports',
            'icon': '✅',
            'description': 'Compliance monitoring and enforcement',
            'reports': [
                {'id': 'compliance_status', 'name': 'Compliance Status Summary', 'description': 'Overall compliance overview'},
                {'id': 'non_filers', 'name': 'Non-Filer Report', 'description': 'Non-filing taxpayers'},
                {'id': 'late_filers', 'name': 'Late Filer Report', 'description': 'Late filing analysis'},
                {'id': 'payment_defaults', 'name': 'Payment Default Report', 'description': 'Payment default tracking'},
                {'id': 'compliance_monitoring', 'name': 'Compliance Monitoring Report', 'description': 'Monitoring activities'},
                {'id': 'enforcement_recovery', 'name': 'Enforcement/Recovery Report', 'description': 'Recovery status'},
            ]
        },
        {
            'id': 'risk',
            'name': 'Risk & Selection Reports',
            'icon': '⚠️',
            'description': 'Risk assessment and selection analysis',
            'reports': [
                {'id': 'risk_summary', 'name': 'Risk Summary', 'description': 'Risk level distribution'},
                {'id': 'risk_type_summary', 'name': 'Risk Type Summary', 'description': 'Risk type analysis'},
                {'id': 'risk_indicator', 'name': 'Risk Indicator Report', 'description': 'Risk indicator breakdown'},
                {'id': 'risk_selection', 'name': 'Risk Selection Report', 'description': 'Selection decisions'},
                {'id': 'selection_summary', 'name': 'Selection Summary', 'description': 'Final selection overview'},
            ]
        },
        {
            'id': 'audit',
            'name': 'Audit Reports',
            'icon': '🔍',
            'description': 'Audit case management and outcomes',
            'reports': [
                {'id': 'audit_case_register', 'name': 'Audit Case Register', 'description': 'Complete audit case list'},
                {'id': 'audit_status', 'name': 'Audit Status Summary', 'description': 'Audit status overview'},
                {'id': 'audit_workload', 'name': 'Audit Officer Workload', 'description': 'Officer workload analysis'},
                {'id': 'audit_findings', 'name': 'Audit Findings Report', 'description': 'Audit findings summary'},
                {'id': 'audit_outcome', 'name': 'Audit Outcome Report', 'description': 'Audit outcome analysis'},
                {'id': 'audit_assessment', 'name': 'Audit Assessment Report', 'description': 'Assessment comparisons'},
                {'id': 'audit_revenue', 'name': 'Audit Revenue Impact', 'description': 'Revenue impact analysis'},
            ]
        },
        {
            'id': 'refund',
            'name': 'Refund Reports',
            'icon': '💸',
            'description': 'Refund processing and analysis',
            'reports': [
                {'id': 'refund_register', 'name': 'Refund Register', 'description': 'Complete refund register'},
                {'id': 'refund_status', 'name': 'Refund Status Summary', 'description': 'Refund status overview'},
                {'id': 'refund_amount', 'name': 'Refund Amount Summary', 'description': 'Refund amount analysis'},
                {'id': 'refund_processing', 'name': 'Refund Processing Time', 'description': 'Processing time analysis'},
                {'id': 'refund_by_period', 'name': 'Refund by Tax Period', 'description': 'Period-wise refund analysis'},
            ]
        },
        {
            'id': 'enforcement',
            'name': 'Enforcement Reports',
            'icon': '⚖️',
            'description': 'Enforcement and recovery tracking',
            'reports': [
                {'id': 'enforcement_register', 'name': 'Enforcement Case Register', 'description': 'Complete enforcement register'},
                {'id': 'outstanding_recovery', 'name': 'Outstanding Recovery Report', 'description': 'Outstanding amounts'},
                {'id': 'recovery_summary', 'name': 'Recovery Summary', 'description': 'Recovery overview'},
                {'id': 'non_payment', 'name': 'Non-Payment Case Report', 'description': 'Non-payment cases'},
                {'id': 'enforcement_status', 'name': 'Enforcement Status Report', 'description': 'Enforcement status overview'},
            ]
        },
        {
            'id': 'officer',
            'name': 'Officer / Workload Reports',
            'icon': '👔',
            'description': 'Officer performance and workload',
            'reports': [
                {'id': 'officer_workload', 'name': 'Officer Workload Report', 'description': 'Complete workload analysis'},
                {'id': 'audit_performance', 'name': 'Audit Performance', 'description': 'Audit performance metrics'},
                {'id': 'refund_performance', 'name': 'Refund Processing Performance', 'description': 'Refund processing metrics'},
            ]
        },
        {
            'id': 'custom',
            'name': 'Custom Reports',
            'icon': '🔧',
            'description': 'Build custom reports',
            'reports': [
                {'id': 'custom_builder', 'name': 'Custom Report Builder', 'description': 'Create custom reports'},
            ]
        },
    ]
    
    context = {
        'report_categories': report_categories,
        'title': 'GST Reports',
    }
    
    return render(request, 'reporting/report_landing.html', context)


@staff_member_required
def report_view(request, report_id):
    """Generic report view handler"""
    engine = ReportEngine()
    
    # Route to specific report handler based on report_id
    report_handlers = {
        'executive_summary': executive_summary_report,
        'taxpayer_master': taxpayer_master_report,
        'new_registrations': new_registration_report,
        'deregistrations': deregistration_report,
        'taxpayer_by_dzongkhag': taxpayer_by_dzongkhag_report,
        'taxpayer_by_sector': taxpayer_by_sector_report,
        'taxpayer_by_org_type': taxpayer_by_org_type_report,
        'taxpayer_by_frequency': taxpayer_by_frequency_report,
        'return_filing_summary': return_filing_summary_report,
        'return_filing_detail': return_filing_detail_report,
        'gst_declaration_summary': gst_declaration_summary_report,
        'payment_compliance': payment_compliance_report,
        'negative_returns': negative_return_report,
        'return_amendments': return_amendment_report,
        'compliance_status': compliance_status_report,
        'non_filers': non_filer_report,
        'late_filers': late_filer_report,
        'payment_defaults': payment_default_report,
        'compliance_monitoring': compliance_monitoring_report,
        'enforcement_recovery': enforcement_recovery_report,
        'risk_summary': risk_summary_report,
        'risk_type_summary': risk_type_summary_report,
        'risk_indicator': risk_indicator_report,
        'risk_selection': risk_selection_report,
        'selection_summary': selection_summary_report,
        'audit_case_register': audit_case_register_report,
        'audit_status': audit_status_report,
        'audit_workload': audit_workload_report,
        'audit_findings': audit_findings_report,
        'audit_outcome': audit_outcome_report,
        'audit_assessment': audit_assessment_report,
        'audit_revenue': audit_revenue_report,
        'refund_register': refund_register_report,
        'refund_status': refund_status_report,
        'refund_amount': refund_amount_report,
        'refund_processing': refund_processing_report,
        'refund_by_period': refund_by_period_report,
        'enforcement_register': enforcement_register_report,
        'outstanding_recovery': outstanding_recovery_report,
        'recovery_summary': recovery_summary_report,
        'non_payment': non_payment_report,
        'enforcement_status': enforcement_status_report,
        'officer_workload': officer_workload_report,
        'audit_performance': audit_performance_report,
        'refund_performance': refund_performance_report,
        'custom_builder': custom_report_builder,
    }
    
    handler = report_handlers.get(report_id)
    if handler:
        return handler(request, engine)
    else:
        messages.error(request, 'Report not found')
        return redirect('report_landing')


# Management Reports Implementation
@staff_member_required
def executive_summary_report(request, engine):
    """Executive GST Summary - Management Report"""
    
    # Get filter parameters
    from_period = request.GET.get('from_tax_period')
    to_period = request.GET.get('to_tax_period')
    
    # Build base querysets
    returns_queryset = GSTReturn.objects.all()
    risk_queryset = ComplianceRiskReferral.objects.all()
    audit_queryset = AuditCase.objects.all()
    refund_queryset = RefundRegister.objects.all()
    enforcement_queryset = EnforcementRecovery.objects.all()
    
    # Apply period filters
    if from_period:
        returns_queryset = returns_queryset.filter(tax_period__gte=from_period)
        risk_queryset = risk_queryset.filter(assessment_from_period__gte=from_period)
        audit_queryset = audit_queryset.filter(from_tax_period__gte=from_period)
        refund_queryset = refund_queryset.filter(tax_period__gte=from_period)
        enforcement_queryset = enforcement_queryset.filter(tax_period__gte=from_period)
    
    if to_period:
        returns_queryset = returns_queryset.filter(tax_period__lte=to_period)
        risk_queryset = risk_queryset.filter(assessment_to_period__lte=to_period)
        audit_queryset = audit_queryset.filter(to_tax_period__lte=to_period)
        refund_queryset = refund_queryset.filter(tax_period__lte=to_period)
        enforcement_queryset = enforcement_queryset.filter(tax_period__lte=to_period)
    
    # Calculate KPIs
    kpis = {
        'taxpayers': {
            'active': TaxpayerMaster.objects.filter(status='Active').count(),
            'new_registrations': TaxpayerMaster.objects.filter(
                registration_date__gte=timezone.now() - timedelta(days=90)
            ).count(),
            'deregistrations': TaxpayerMaster.objects.filter(status='Deregistered').count(),
        },
        'returns': {
            'expected': returns_queryset.count(),
            'filed': returns_queryset.filter(filing_status='Filed').count(),
            'late': returns_queryset.filter(filing_status='Late Filed').count(),
            'not_filed': returns_queryset.filter(filing_status='Not Filed').count(),
        },
        'gst': {
            'declared_sales': engine.format_number(returns_queryset.aggregate(Sum('declared_sales'))['declared_sales__sum'] or 0),
            'output_gst': engine.format_number(returns_queryset.aggregate(Sum('declared_output_gst'))['declared_output_gst__sum'] or 0),
            'itc_claimed': engine.format_number(returns_queryset.aggregate(Sum('total_itc_claimed'))['total_itc_claimed__sum'] or 0),
            'gst_payable': engine.format_number(returns_queryset.aggregate(Sum('gst_payable_refundable'))['gst_payable_refundable__sum'] or 0),
            'actual_payment': engine.format_number(returns_queryset.aggregate(Sum('actual_gst_payment_received'))['actual_gst_payment_received__sum'] or 0),
        },
        'compliance': {
            'compliant': returns_queryset.filter(compliance_status='Compliant').count(),
            'late_filers': returns_queryset.filter(compliance_status='Late Filer').count(),
            'non_filers': returns_queryset.filter(compliance_status='Non-Filer').count(),
            'payment_defaults': returns_queryset.filter(payment_status='Default').count(),
        },
        'risk': {
            'low': risk_queryset.filter(risk_level='Low').count(),
            'medium': risk_queryset.filter(risk_level='Medium').count(),
            'high': risk_queryset.filter(risk_level='High').count(),
            'critical': risk_queryset.filter(risk_level='Critical').count(),
        },
        'selection': {
            'audit': risk_queryset.filter(final_selection='AUDIT').count(),
            'review': risk_queryset.filter(final_selection='REVIEW').count(),
            'monitor': risk_queryset.filter(final_selection='MONITOR').count(),
            'not_selected': risk_queryset.filter(final_selection='NOT SELECTED').count(),
        },
        'audit': {
            'total': audit_queryset.count(),
            'pending': audit_queryset.filter(status='Pending Assignment').count(),
            'in_progress': audit_queryset.filter(status='In Progress').count(),
            'completed': audit_queryset.filter(status='Completed').count(),
            'closed': audit_queryset.filter(status='Closed').count(),
        },
        'refund': {
            'total': refund_queryset.count(),
            'claimed': engine.format_number(refund_queryset.aggregate(Sum('claimed_amount'))['claimed_amount__sum'] or 0),
            'approved': engine.format_number(refund_queryset.aggregate(Sum('refund_approved'))['refund_approved__sum'] or 0),
            'adjusted': engine.format_number(refund_queryset.aggregate(Sum('adjustment'))['adjustment__sum'] or 0),
        },
        'enforcement': {
            'cases': enforcement_queryset.count(),
            'amount_due': engine.format_number(enforcement_queryset.aggregate(Sum('amount_due'))['amount_due__sum'] or 0),
            'amount_recovered': engine.format_number(enforcement_queryset.aggregate(Sum('amount_recovered'))['amount_recovered__sum'] or 0),
        },
    }
    
    # Calculate compliance percentages
    kpis['returns']['filing_compliance'] = engine.calculate_percentage(
        kpis['returns']['filed'], kpis['returns']['expected']
    )
    
    context = {
        'report_title': 'Executive GST Summary',
        'report_period': f"{from_period or 'All'} to {to_period or 'All'}",
        'kpis': kpis,
        'filters': {
            'from_tax_period': from_period,
            'to_tax_period': to_period,
            'tax_periods': engine.tax_periods,
        },
        'generated_date': timezone.now().strftime('%d-%m-%Y %H:%M'),
        'generated_by': request.user.username,
    }
    
    return render(request, 'reporting/executive_summary.html', context)


# Taxpayer Reports Implementation
@staff_member_required
def taxpayer_master_report(request, engine):
    """Taxpayer Master Report"""
    
    queryset = TaxpayerMaster.objects.all()
    
    # Apply filters
    filters = {
        'gstin': request.GET.get('gstin'),
        'taxpayer_name': request.GET.get('taxpayer_name'),
        'dzongkhag': request.GET.get('dzongkhag'),
        'sector': request.GET.get('sector'),
        'organisation_type': request.GET.get('organisation_type'),
        'status': request.GET.get('status'),
    }
    
    queryset = engine.apply_filters(queryset, filters)
    
    # Pagination
    page = request.GET.get('page', 1)
    paginator = Paginator(queryset, 50)
    taxpayers = paginator.get_page(page)
    
    context = {
        'report_title': 'Taxpayer Master Report',
        'taxpayers': taxpayers,
        'filters': filters,
        'filter_options': {
            'dzongkhags': engine.dzongkhags,
            'sectors': engine.sectors,
            'organisation_types': engine.organisation_types,
            'statuses': ['Active', 'Inactive', 'Deregistered', 'Suspended'],
        },
        'generated_date': timezone.now().strftime('%d-%m-%Y %H:%M'),
        'generated_by': request.user.username,
    }
    
    return render(request, 'reporting/taxpayer_master.html', context)


@staff_member_required
def new_registration_report(request, engine):
    """New Registration Report"""
    
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    dzongkhag = request.GET.get('dzongkhag')
    sector = request.GET.get('sector')
    organisation_type = request.GET.get('organisation_type')
    
    queryset = TaxpayerMaster.objects.all()
    
    if from_date:
        queryset = queryset.filter(registration_date__gte=from_date)
    if to_date:
        queryset = queryset.filter(registration_date__lte=to_date)
    if dzongkhag:
        queryset = queryset.filter(dzongkhag=dzongkhag)
    if sector:
        queryset = queryset.filter(sector=sector)
    if organisation_type:
        queryset = queryset.filter(organisation_type=organisation_type)
    
    # Pagination
    page = request.GET.get('page', 1)
    paginator = Paginator(queryset, 50)
    taxpayers = paginator.get_page(page)
    
    context = {
        'report_title': 'New Registration Report',
        'taxpayers': taxpayers,
        'filters': {
            'from_date': from_date,
            'to_date': to_date,
            'dzongkhag': dzongkhag,
            'sector': sector,
            'organisation_type': organisation_type,
        },
        'filter_options': {
            'dzongkhags': engine.dzongkhags,
            'sectors': engine.sectors,
            'organisation_types': engine.organisation_types,
        },
        'generated_date': timezone.now().strftime('%d-%m-%Y %H:%M'),
        'generated_by': request.user.username,
    }
    
    return render(request, 'reporting/new_registrations.html', context)


# Placeholder functions for other reports (to be implemented)
@staff_member_required
def deregistration_report(request, engine):
    """Deregistration Report"""
    queryset = TaxpayerMaster.objects.filter(status='Deregistered')
    
    # Apply filters
    filters = {
        'gstin': request.GET.get('gstin'),
        'taxpayer_name': request.GET.get('taxpayer_name'),
        'dzongkhag': request.GET.get('dzongkhag'),
        'sector': request.GET.get('sector'),
        'organisation_type': request.GET.get('organisation_type'),
    }
    
    queryset = engine.apply_filters(queryset, filters)
    
    context = {
        'report_title': 'Deregistration Report',
        'taxpayers': queryset,
        'filters': filters,
        'filter_options': {
            'dzongkhags': engine.dzongkhags,
            'sectors': engine.sectors,
            'organisation_types': engine.organisation_types,
        },
        'generated_date': timezone.now().strftime('%d-%m-%Y %H:%M'),
        'generated_by': request.user.username,
    }
    
    return render(request, 'reporting/deregistrations.html', context)


@staff_member_required
def taxpayer_by_dzongkhag_report(request, engine):
    """Taxpayer by Dzongkhag Report"""
    data = TaxpayerMaster.objects.values('dzongkhag').annotate(
        count=Count('id')
    ).order_by('-count')
    
    total = TaxpayerMaster.objects.count()
    
    for item in data:
        item['percentage'] = engine.calculate_percentage(item['count'], total)
    
    context = {
        'report_title': 'Taxpayer by Dzongkhag',
        'data': data,
        'total': total,
        'generated_date': timezone.now().strftime('%d-%m-%Y %H:%M'),
        'generated_by': request.user.username,
    }
    
    return render(request, 'reporting/taxpayer_by_dzongkhag.html', context)


@staff_member_required
def taxpayer_by_sector_report(request, engine):
    """Taxpayer by Sector Report"""
    data = TaxpayerMaster.objects.values('sector').annotate(
        count=Count('id')
    ).order_by('-count')
    
    total = TaxpayerMaster.objects.count()
    
    for item in data:
        item['percentage'] = engine.calculate_percentage(item['count'], total)
    
    context = {
        'report_title': 'Taxpayer by Sector',
        'data': data,
        'total': total,
        'generated_date': timezone.now().strftime('%d-%m-%Y %H:%M'),
        'generated_by': request.user.username,
    }
    
    return render(request, 'reporting/taxpayer_by_sector.html', context)


@staff_member_required
def taxpayer_by_org_type_report(request, engine):
    """Taxpayer by Organisation Type Report"""
    data = TaxpayerMaster.objects.values('organisation_type').annotate(
        count=Count('id')
    ).order_by('-count')
    
    total = TaxpayerMaster.objects.count()
    
    for item in data:
        item['percentage'] = engine.calculate_percentage(item['count'], total)
    
    context = {
        'report_title': 'Taxpayer by Organisation Type',
        'data': data,
        'total': total,
        'generated_date': timezone.now().strftime('%d-%m-%Y %H:%M'),
        'generated_by': request.user.username,
    }
    
    return render(request, 'reporting/taxpayer_by_org_type.html', context)


@staff_member_required
def taxpayer_by_frequency_report(request, engine):
    """Taxpayer by Filing Frequency Report"""
    data = TaxpayerMaster.objects.values('frequency').annotate(
        count=Count('id')
    ).order_by('-count')
    
    total = TaxpayerMaster.objects.count()
    
    for item in data:
        item['percentage'] = engine.calculate_percentage(item['count'], total)
    
    context = {
        'report_title': 'Taxpayer by Filing Frequency',
        'data': data,
        'total': total,
        'generated_date': timezone.now().strftime('%d-%m-%Y %H:%M'),
        'generated_by': request.user.username,
    }
    
    return render(request, 'reporting/taxpayer_by_frequency.html', context)


# Placeholder functions for all other reports (simplified for brevity)
@staff_member_required
def return_filing_summary_report(request, engine):
    """Return Filing Summary Report"""
    from_period = request.GET.get('from_tax_period')
    to_period = request.GET.get('to_tax_period')
    
    queryset = GSTReturn.objects.all()
    if from_period:
        queryset = queryset.filter(tax_period__gte=from_period)
    if to_period:
        queryset = queryset.filter(tax_period__lte=to_period)
    
    summary = {
        'expected': queryset.count(),
        'filed': queryset.filter(filing_status='Filed').count(),
        'late_filed': queryset.filter(filing_status='Late Filed').count(),
        'not_filed': queryset.filter(filing_status='Not Filed').count(),
    }
    
    summary['filing_compliance'] = engine.calculate_percentage(summary['filed'], summary['expected'])
    
    context = {
        'report_title': 'Return Filing Summary',
        'summary': summary,
        'report_period': f"{from_period or 'All'} to {to_period or 'All'}",
        'filters': {
            'from_tax_period': from_period,
            'to_tax_period': to_period,
            'tax_periods': engine.tax_periods,
        },
        'generated_date': timezone.now().strftime('%d-%m-%Y %H:%M'),
        'generated_by': request.user.username,
    }
    
    return render(request, 'reporting/return_filing_summary.html', context)


@staff_member_required
def return_filing_detail_report(request, engine):
    """Return Filing Detail Report"""
    from_period = request.GET.get('from_tax_period')
    to_period = request.GET.get('to_tax_period')
    filing_status = request.GET.get('filing_status')
    payment_status = request.GET.get('payment_status')
    gstin = request.GET.get('gstin')
    taxpayer_name = request.GET.get('taxpayer_name')
    
    queryset = GSTReturn.objects.all()
    
    if from_period:
        queryset = queryset.filter(tax_period__gte=from_period)
    if to_period:
        queryset = queryset.filter(tax_period__lte=to_period)
    if filing_status:
        queryset = queryset.filter(filing_status=filing_status)
    if payment_status:
        queryset = queryset.filter(payment_status=payment_status)
    if gstin:
        queryset = queryset.filter(gstin__icontains=gstin)
    if taxpayer_name:
        queryset = queryset.filter(taxpayer_name__icontains=taxpayer_name)
    
    # Calculate filing delay
    returns_list = []
    for return_obj in queryset:
        filing_delay = None
        if return_obj.return_filing_date and return_obj.return_due_date:
            filing_delay = (return_obj.return_filing_date - return_obj.return_due_date).days
        
        returns_list.append({
            'tax_period': return_obj.tax_period,
            'gstin': return_obj.gstin,
            'taxpayer_name': return_obj.taxpayer_name,
            'due_date': return_obj.return_due_date,
            'filing_date': return_obj.return_filing_date,
            'filing_delay': filing_delay,
            'filing_status': return_obj.filing_status,
            'payment_status': return_obj.payment_status,
            'compliance_status': return_obj.compliance_status,
            'id': return_obj.id,
        })
    
    # Pagination
    page = request.GET.get('page', 1)
    paginator = Paginator(returns_list, 50)
    returns = paginator.get_page(page)
    
    context = {
        'report_title': 'Return Filing Detail',
        'returns': returns,
        'filters': {
            'from_tax_period': from_period,
            'to_tax_period': to_period,
            'filing_status': filing_status,
            'payment_status': payment_status,
            'gstin': gstin,
            'taxpayer_name': taxpayer_name,
        },
        'filter_options': {
            'tax_periods': engine.tax_periods,
            'filing_statuses': ['Filed', 'Late Filed', 'Not Filed', 'Amended'],
            'payment_statuses': ['Paid', 'Partial', 'Default', 'No Liability'],
        },
        'generated_date': timezone.now().strftime('%d-%m-%Y %H:%M'),
        'generated_by': request.user.username,
    }
    
    return render(request, 'reporting/return_filing_detail.html', context)


@staff_member_required
def gst_declaration_summary_report(request, engine):
    """GST Declaration Summary Report"""
    from_period = request.GET.get('from_tax_period')
    to_period = request.GET.get('to_tax_period')
    
    queryset = GSTReturn.objects.all()
    if from_period:
        queryset = queryset.filter(tax_period__gte=from_period)
    if to_period:
        queryset = queryset.filter(tax_period__lte=to_period)
    
    summary = queryset.aggregate(
        declared_sales=Sum('declared_sales'),
        domestic_purchases=Sum('declared_domestic_purchase'),
        import_value=Sum('declared_import_value'),
        output_gst=Sum('declared_output_gst'),
        itc=Sum('total_itc_claimed'),
        gst_payable=Sum('gst_payable_refundable'),
        actual_payment=Sum('actual_gst_payment_received')
    )
    
    context = {
        'report_title': 'GST Declaration Summary',
        'summary': {
            'declared_sales': engine.format_number(summary['declared_sales'] or 0),
            'domestic_purchases': engine.format_number(summary['domestic_purchases'] or 0),
            'import_value': engine.format_number(summary['import_value'] or 0),
            'output_gst': engine.format_number(summary['output_gst'] or 0),
            'itc': engine.format_number(summary['itc'] or 0),
            'gst_payable': engine.format_number(summary['gst_payable'] or 0),
            'actual_payment': engine.format_number(summary['actual_payment'] or 0),
        },
        'report_period': f"{from_period or 'All'} to {to_period or 'All'}",
        'filters': {
            'from_tax_period': from_period,
            'to_tax_period': to_period,
            'tax_periods': engine.tax_periods,
        },
        'generated_date': timezone.now().strftime('%d-%m-%Y %H:%M'),
        'generated_by': request.user.username,
    }
    
    return render(request, 'reporting/gst_declaration_summary.html', context)


@staff_member_required
def payment_compliance_report(request, engine):
    """Payment Compliance Report"""
    from_period = request.GET.get('from_tax_period')
    to_period = request.GET.get('to_tax_period')
    
    queryset = GSTReturn.objects.all()
    if from_period:
        queryset = queryset.filter(tax_period__gte=from_period)
    if to_period:
        queryset = queryset.filter(tax_period__lte=to_period)
    
    # Calculate payment compliance
    total_with_liability = queryset.filter(gst_payable_refundable__gt=0).count()
    total_paid = queryset.filter(gst_payable_refundable__gt=0, payment_status='Paid').count()
    
    payment_compliance = engine.calculate_percentage(total_paid, total_with_liability)
    
    payment_details = []
    for return_obj in queryset:
        gst_payable = return_obj.gst_payable_refundable or 0
        actual_payment = return_obj.actual_gst_payment_received or 0
        outstanding = gst_payable - actual_payment if gst_payable > 0 else 0
        
        payment_details.append({
            'tax_period': return_obj.tax_period,
            'gstin': return_obj.gstin,
            'taxpayer_name': return_obj.taxpayer_name,
            'gst_payable': engine.format_number(gst_payable),
            'actual_payment': engine.format_number(actual_payment),
            'outstanding': engine.format_number(outstanding),
            'payment_status': return_obj.payment_status,
            'id': return_obj.id,
        })
    
    # Pagination
    page = request.GET.get('page', 1)
    paginator = Paginator(payment_details, 50)
    payments = paginator.get_page(page)
    
    context = {
        'report_title': 'Payment Compliance Report',
        'compliance_summary': {
            'total_with_liability': total_with_liability,
            'total_paid': total_paid,
            'payment_compliance': payment_compliance,
        },
        'payments': payments,
        'report_period': f"{from_period or 'All'} to {to_period or 'All'}",
        'filters': {
            'from_tax_period': from_period,
            'to_tax_period': to_period,
            'tax_periods': engine.tax_periods,
        },
        'generated_date': timezone.now().strftime('%d-%m-%Y %H:%M'),
        'generated_by': request.user.username,
    }
    
    return render(request, 'reporting/payment_compliance.html', context)


@staff_member_required
def negative_return_report(request, engine):
    """Negative Return Report"""
    from_period = request.GET.get('from_tax_period')
    to_period = request.GET.get('to_tax_period')
    gstin = request.GET.get('gstin')
    
    queryset = GSTReturn.objects.filter(gst_payable_refundable__lt=0)
    
    if from_period:
        queryset = queryset.filter(tax_period__gte=from_period)
    if to_period:
        queryset = queryset.filter(tax_period__lte=to_period)
    if gstin:
        queryset = queryset.filter(gstin__icontains=gstin)
    
    negative_returns = []
    for return_obj in queryset:
        negative_returns.append({
            'tax_period': return_obj.tax_period,
            'gstin': return_obj.gstin,
            'taxpayer_name': return_obj.taxpayer_name,
            'output_gst': engine.format_number(return_obj.declared_output_gst or 0),
            'itc': engine.format_number(return_obj.total_itc_claimed or 0),
            'gst_payable': engine.format_number(return_obj.gst_payable_refundable or 0),
            'id': return_obj.id,
        })
    
    # Pagination
    page = request.GET.get('page', 1)
    paginator = Paginator(negative_returns, 50)
    returns = paginator.get_page(page)
    
    context = {
        'report_title': 'Negative Return Report',
        'returns': returns,
        'report_period': f"{from_period or 'All'} to {to_period or 'All'}",
        'filters': {
            'from_tax_period': from_period,
            'to_tax_period': to_period,
            'gstin': gstin,
            'tax_periods': engine.tax_periods,
        },
        'generated_date': timezone.now().strftime('%d-%m-%Y %H:%M'),
        'generated_by': request.user.username,
    }
    
    return render(request, 'reporting/negative_return.html', context)


@staff_member_required
def return_amendment_report(request, engine):
    """Return Amendment Report"""
    from_period = request.GET.get('from_tax_period')
    to_period = request.GET.get('to_tax_period')
    gstin = request.GET.get('gstin')
    
    # Filter for amended returns
    queryset = GSTReturn.objects.filter(filing_status='Amended')
    
    if from_period:
        queryset = queryset.filter(tax_period__gte=from_period)
    if to_period:
        queryset = queryset.filter(tax_period__lte=to_period)
    if gstin:
        queryset = queryset.filter(gstin__icontains=gstin)
    
    amendments = []
    for return_obj in queryset:
        amendments.append({
            'tax_period': return_obj.tax_period,
            'gstin': return_obj.gstin,
            'taxpayer_name': return_obj.taxpayer_name,
            'amendment_date': return_obj.return_filing_date,
            'original_value': engine.format_number(return_obj.declared_sales or 0),  # Using declared sales as example
            'amended_value': engine.format_number(return_obj.declared_sales or 0),  # Would need original vs amended comparison
            'variation': 0,  # Would need amendment logic
            'reason': return_obj.remarks or 'Not specified',
            'id': return_obj.id,
        })
    
    # Pagination
    page = request.GET.get('page', 1)
    paginator = Paginator(amendments, 50)
    amendments_list = paginator.get_page(page)
    
    context = {
        'report_title': 'Return Amendment Report',
        'amendments': amendments_list,
        'report_period': f"{from_period or 'All'} to {to_period or 'All'}",
        'filters': {
            'from_tax_period': from_period,
            'to_tax_period': to_period,
            'gstin': gstin,
            'tax_periods': engine.tax_periods,
        },
        'generated_date': timezone.now().strftime('%d-%m-%Y %H:%M'),
        'generated_by': request.user.username,
    }
    
    return render(request, 'reporting/return_amendment.html', context)

@staff_member_required
def return_amendment_report(request, engine):
    """Return Amendment Report"""
    from_period = request.GET.get('from_tax_period')
    to_period = request.GET.get('to_tax_period')
    gstin = request.GET.get('gstin')
    
    # Filter for amended returns
    queryset = GSTReturn.objects.filter(filing_status='Amended')
    
    if from_period:
        queryset = queryset.filter(tax_period__gte=from_period)
    if to_period:
        queryset = queryset.filter(tax_period__lte=to_period)
    if gstin:
        queryset = queryset.filter(gstin__icontains=gstin)
    
    amendments = []
    for return_obj in queryset:
        amendments.append({
            'tax_period': return_obj.tax_period,
            'gstin': return_obj.gstin,
            'taxpayer_name': return_obj.taxpayer_name,
            'amendment_date': return_obj.return_filing_date,
            'original_value': engine.format_number(return_obj.declared_sales or 0),  # Using declared sales as example
            'amended_value': engine.format_number(return_obj.declared_sales or 0),  # Would need original vs amended comparison
            'variation': 0,  # Would need amendment logic
            'reason': return_obj.remarks or 'Not specified',
            'id': return_obj.id,
        })
    
    # Pagination
    page = request.GET.get('page', 1)
    paginator = Paginator(amendments, 50)
    amendments_list = paginator.get_page(page)
    
    context = {
        'report_title': 'Return Amendment Report',
        'amendments': amendments_list,
        'report_period': f"{from_period or 'All'} to {to_period or 'All'}",
        'filters': {
            'from_tax_period': from_period,
            'to_tax_period': to_period,
            'gstin': gstin,
            'tax_periods': engine.tax_periods,
        },
        'generated_date': timezone.now().strftime('%d-%m-%Y %H:%M'),
        'generated_by': request.user.username,
    }
    
    return render(request, 'reporting/return_amendment.html', context)


@staff_member_required
def compliance_status_report(request, engine):
    """Compliance Status Summary Report"""
    from_period = request.GET.get('from_tax_period')
    to_period = request.GET.get('to_tax_period')
    
    queryset = GSTReturn.objects.all()
    if from_period:
        queryset = queryset.filter(tax_period__gte=from_period)
    if to_period:
        queryset = queryset.filter(tax_period__lte=to_period)
    
    summary = {
        'compliant': queryset.filter(compliance_status='Compliant').count(),
        'late_filer': queryset.filter(compliance_status='Late Filer').count(),
        'non_filer': queryset.filter(compliance_status='Non-Filer').count(),
        'late_payment': queryset.filter(compliance_status='Late Payment').count(),
        'payment_default': queryset.filter(compliance_status='Payment Default').count(),
        'return_amendment': queryset.filter(compliance_status='Return Amendment').count(),
    }
    
    context = {
        'report_title': 'Compliance Status Summary',
        'summary': summary,
        'report_period': f"{from_period or 'All'} to {to_period or 'All'}",
        'filters': {
            'from_tax_period': from_period,
            'to_tax_period': to_period,
            'tax_periods': engine.tax_periods,
        },
        'generated_date': timezone.now().strftime('%d-%m-%Y %H:%M'),
        'generated_by': request.user.username,
    }
    
    return render(request, 'reporting/compliance_status.html', context)


@staff_member_required
def non_filer_report(request, engine):
    """Non-Filer Report"""
    from_period = request.GET.get('from_tax_period')
    to_period = request.GET.get('to_tax_period')
    
    queryset = GSTReturn.objects.filter(filing_status='Not Filed')
    if from_period:
        queryset = queryset.filter(tax_period__gte=from_period)
    if to_period:
        queryset = queryset.filter(tax_period__lte=to_period)
    
    non_filers = []
    for return_obj in queryset:
        non_filers.append({
            'tax_period': return_obj.tax_period,
            'gstin': return_obj.gstin,
            'taxpayer_name': return_obj.taxpayer_name,
            'due_date': return_obj.return_due_date,
            'filing_status': return_obj.filing_status,
            'payment_status': return_obj.payment_status,
            'compliance_status': return_obj.compliance_status,
            'id': return_obj.id,
        })
    
    # Pagination
    page = request.GET.get('page', 1)
    paginator = Paginator(non_filers, 50)
    non_filers_list = paginator.get_page(page)
    
    context = {
        'report_title': 'Non-Filer Report',
        'non_filers': non_filers_list,
        'report_period': f"{from_period or 'All'} to {to_period or 'All'}",
        'filters': {
            'from_tax_period': from_period,
            'to_tax_period': to_period,
            'tax_periods': engine.tax_periods,
        },
        'generated_date': timezone.now().strftime('%d-%m-%Y %H:%M'),
        'generated_by': request.user.username,
    }
    
    return render(request, 'reporting/non_filer.html', context)


@staff_member_required
def late_filer_report(request, engine):
    """Late Filer Report"""
    from_period = request.GET.get('from_tax_period')
    to_period = request.GET.get('to_tax_period')
    
    queryset = GSTReturn.objects.filter(filing_status='Late Filed')
    if from_period:
        queryset = queryset.filter(tax_period__gte=from_period)
    if to_period:
        queryset = queryset.filter(tax_period__lte=to_period)
    
    late_filers = []
    for return_obj in queryset:
        filing_delay = None
        if return_obj.return_filing_date and return_obj.return_due_date:
            filing_delay = (return_obj.return_filing_date - return_obj.return_due_date).days
        
        late_filers.append({
            'tax_period': return_obj.tax_period,
            'gstin': return_obj.gstin,
            'taxpayer_name': return_obj.taxpayer_name,
            'due_date': return_obj.return_due_date,
            'filing_date': return_obj.return_filing_date,
            'filing_delay': filing_delay,
            'payment_status': return_obj.payment_status,
            'compliance_status': return_obj.compliance_status,
            'id': return_obj.id,
        })
    
    # Pagination
    page = request.GET.get('page', 1)
    paginator = Paginator(late_filers, 50)
    late_filers_list = paginator.get_page(page)
    
    context = {
        'report_title': 'Late Filer Report',
        'late_filers': late_filers_list,
        'report_period': f"{from_period or 'All'} to {to_period or 'All'}",
        'filters': {
            'from_tax_period': from_period,
            'to_tax_period': to_period,
            'tax_periods': engine.tax_periods,
        },
        'generated_date': timezone.now().strftime('%d-%m-%Y %H:%M'),
        'generated_by': request.user.username,
    }
    
    return render(request, 'reporting/late_filer.html', context)


@staff_member_required
def payment_default_report(request, engine):
    """Payment Default Report"""
    from_period = request.GET.get('from_tax_period')
    to_period = request.GET.get('to_tax_period')
    
    queryset = GSTReturn.objects.filter(payment_status='Default')
    if from_period:
        queryset = queryset.filter(tax_period__gte=from_period)
    if to_period:
        queryset = queryset.filter(tax_period__lte=to_period)
    
    payment_defaults = []
    for return_obj in queryset:
        gst_payable = return_obj.gst_payable_refundable or 0
        actual_payment = return_obj.actual_gst_payment_received or 0
        outstanding = gst_payable - actual_payment if gst_payable > 0 else 0
        
        payment_defaults.append({
            'tax_period': return_obj.tax_period,
            'gstin': return_obj.gstin,
            'taxpayer_name': return_obj.taxpayer_name,
            'gst_payable': engine.format_number(gst_payable),
            'actual_payment': engine.format_number(actual_payment),
            'outstanding': engine.format_number(outstanding),
            'payment_status': return_obj.payment_status,
            'compliance_status': return_obj.compliance_status,
            'id': return_obj.id,
        })
    
    # Pagination
    page = request.GET.get('page', 1)
    paginator = Paginator(payment_defaults, 50)
    payment_defaults_list = paginator.get_page(page)
    
    context = {
        'report_title': 'Payment Default Report',
        'payment_defaults': payment_defaults_list,
        'report_period': f"{from_period or 'All'} to {to_period or 'All'}",
        'filters': {
            'from_tax_period': from_period,
            'to_tax_period': to_period,
            'tax_periods': engine.tax_periods,
        },
        'generated_date': timezone.now().strftime('%d-%m-%Y %H:%M'),
        'generated_by': request.user.username,
    }
    
    return render(request, 'reporting/payment_default.html', context)


@staff_member_required
def compliance_monitoring_report(request, engine):
    """Compliance Monitoring Report"""
    queryset = ComplianceMonitoring.objects.all()
    
    # Apply filters
    filters = {
        'gstin': request.GET.get('gstin'),
        'taxpayer_name': request.GET.get('taxpayer_name'),
        'tax_period': request.GET.get('tax_period'),
        'compliance_status': request.GET.get('compliance_status'),
    }
    
    queryset = engine.apply_filters(queryset, filters)
    
    # Pagination
    page = request.GET.get('page', 1)
    paginator = Paginator(queryset, 50)
    monitoring_records = paginator.get_page(page)
    
    context = {
        'report_title': 'Compliance Monitoring Report',
        'monitoring_records': monitoring_records,
        'filters': filters,
        'filter_options': {
            'tax_periods': engine.tax_periods,
            'compliance_statuses': ['Compliant', 'Late Filer', 'Non-Filer', 'Late Payment', 'Payment Default', 'Return Amendment'],
        },
        'generated_date': timezone.now().strftime('%d-%m-%Y %H:%M'),
        'generated_by': request.user.username,
    }
    
    return render(request, 'reporting/compliance_monitoring.html', context)


@staff_member_required
def enforcement_recovery_report(request, engine):
    """Enforcement/Recovery Report"""
    from_period = request.GET.get('from_tax_period')
    to_period = request.GET.get('to_tax_period')
    
    queryset = EnforcementRecovery.objects.all()
    if from_period:
        queryset = queryset.filter(tax_period__gte=from_period)
    if to_period:
        queryset = queryset.filter(tax_period__lte=to_period)
    
    enforcement_cases = []
    for case in queryset:
        outstanding = (case.amount_due or 0) - (case.amount_recovered or 0)
        
        enforcement_cases.append({
            'case_id': case.case_id,
            'gstin': case.gstin,
            'taxpayer_name': case.taxpayer_name,
            'tax_period': case.tax_period,
            'case_type': case.case_type,
            'amount_due': engine.format_number(case.amount_due or 0),
            'amount_recovered': engine.format_number(case.amount_recovered or 0),
            'outstanding': engine.format_number(outstanding),
            'status': case.status,
            'assigned_officer': case.assigned_officer.username if case.assigned_officer else '-',
            'notice_date': case.notice_date,
            'remarks': case.remarks,
            'id': case.id,
        })
    
    # Pagination
    page = request.GET.get('page', 1)
    paginator = Paginator(enforcement_cases, 50)
    enforcement_list = paginator.get_page(page)
    
    context = {
        'report_title': 'Enforcement/Recovery Report',
        'enforcement_cases': enforcement_list,
        'report_period': f"{from_period or 'All'} to {to_period or 'All'}",
        'filters': {
            'from_tax_period': from_period,
            'to_tax_period': to_period,
            'tax_periods': engine.tax_periods,
        },
        'generated_date': timezone.now().strftime('%d-%m-%Y %H:%M'),
        'generated_by': request.user.username,
    }
    
    return render(request, 'reporting/enforcement_recovery.html', context)


@staff_member_required
def risk_summary_report(request, engine):
    """Risk Summary Report"""
    from_period = request.GET.get('from_tax_period')
    to_period = request.GET.get('to_tax_period')
    
    queryset = ComplianceRiskReferral.objects.all()
    if from_period:
        queryset = queryset.filter(assessment_from_period__gte=from_period)
    if to_period:
        queryset = queryset.filter(assessment_to_period__lte=to_period)
    
    summary = {
        'low': queryset.filter(risk_level='Low').count(),
        'medium': queryset.filter(risk_level='Medium').count(),
        'high': queryset.filter(risk_level='High').count(),
        'critical': queryset.filter(risk_level='Critical').count(),
    }
    
    context = {
        'report_title': 'Risk Summary',
        'summary': summary,
        'report_period': f"{from_period or 'All'} to {to_period or 'All'}",
        'filters': {
            'from_tax_period': from_period,
            'to_tax_period': to_period,
            'tax_periods': engine.tax_periods,
        },
        'generated_date': timezone.now().strftime('%d-%m-%Y %H:%M'),
        'generated_by': request.user.username,
    }
    
    return render(request, 'reporting/risk_summary.html', context)


@staff_member_required
def risk_type_summary_report(request, engine):
    """Risk Type Summary Report"""
    from_period = request.GET.get('from_tax_period')
    to_period = request.GET.get('to_tax_period')
    
    queryset = ComplianceRiskReferral.objects.all()
    if from_period:
        queryset = queryset.filter(assessment_from_period__gte=from_period)
    if to_period:
        queryset = queryset.filter(assessment_to_period__lte=to_period)
    
    risk_types = queryset.values('risk_type').annotate(
        count=Count('id')
    ).order_by('-count')
    
    total = queryset.count()
    
    for item in risk_types:
        item['percentage'] = engine.calculate_percentage(item['count'], total)
    
    context = {
        'report_title': 'Risk Type Summary',
        'risk_types': risk_types,
        'total': total,
        'report_period': f"{from_period or 'All'} to {to_period or 'All'}",
        'filters': {
            'from_tax_period': from_period,
            'to_tax_period': to_period,
            'tax_periods': engine.tax_periods,
        },
        'generated_date': timezone.now().strftime('%d-%m-%Y %H:%M'),
        'generated_by': request.user.username,
    }
    
    return render(request, 'reporting/risk_type_summary.html', context)


@staff_member_required
def risk_indicator_report(request, engine):
    """Risk Indicator Report"""
    from_period = request.GET.get('from_tax_period')
    to_period = request.GET.get('to_tax_period')
    
    queryset = ComplianceRiskReferral.objects.all()
    if from_period:
        queryset = queryset.filter(assessment_from_period__gte=from_period)
    if to_period:
        queryset = queryset.filter(assessment_to_period__lte=to_period)
    
    risk_indicators = []
    for risk in queryset:
        risk_indicators.append({
            'risk_id': risk.risk_id,
            'gstin': risk.gstin,
            'taxpayer_name': risk.taxpayer_name,
            'assessment_from_period': risk.assessment_from_period,
            'assessment_to_period': risk.assessment_to_period,
            'risk_level': risk.risk_level,
            'risk_type': risk.risk_type,
            'filing_indicator': risk.filing_indicator,
            'payment_indicator': risk.payment_indicator,
            'assessment_score': risk.assessment_score,
            'final_selection': risk.final_selection,
            'id': risk.id,
        })
    
    # Pagination
    page = request.GET.get('page', 1)
    paginator = Paginator(risk_indicators, 50)
    risk_list = paginator.get_page(page)
    
    context = {
        'report_title': 'Risk Indicator Report',
        'risk_indicators': risk_list,
        'report_period': f"{from_period or 'All'} to {to_period or 'All'}",
        'filters': {
            'from_tax_period': from_period,
            'to_tax_period': to_period,
            'tax_periods': engine.tax_periods,
        },
        'generated_date': timezone.now().strftime('%d-%m-%Y %H:%M'),
        'generated_by': request.user.username,
    }
    
    return render(request, 'reporting/risk_indicator.html', context)


@staff_member_required
def risk_selection_report(request, engine):
    """Risk Selection Report"""
    from_period = request.GET.get('from_tax_period')
    to_period = request.GET.get('to_tax_period')
    
    queryset = ComplianceRiskReferral.objects.all()
    if from_period:
        queryset = queryset.filter(assessment_from_period__gte=from_period)
    if to_period:
        queryset = queryset.filter(assessment_to_period__lte=to_period)
    
    selection_details = []
    for risk in queryset:
        selection_details.append({
            'risk_id': risk.risk_id,
            'gstin': risk.gstin,
            'taxpayer_name': risk.taxpayer_name,
            'assessment_from_period': risk.assessment_from_period,
            'assessment_to_period': risk.assessment_to_period,
            'risk_level': risk.risk_level,
            'risk_type': risk.risk_type,
            'selection_decision': risk.selection_decision,
            'selection_by': risk.selection_by.username if risk.selection_by else '-',
            'selection_date': risk.selection_date,
            'final_selection': risk.final_selection,
            'id': risk.id,
        })
    
    # Pagination
    page = request.GET.get('page', 1)
    paginator = Paginator(selection_details, 50)
    selection_list = paginator.get_page(page)
    
    context = {
        'report_title': 'Risk Selection Report',
        'selection_details': selection_list,
        'report_period': f"{from_period or 'All'} to {to_period or 'All'}",
        'filters': {
            'from_tax_period': from_period,
            'to_tax_period': to_period,
            'tax_periods': engine.tax_periods,
        },
        'generated_date': timezone.now().strftime('%d-%m-%Y %H:%M'),
        'generated_by': request.user.username,
    }
    
    return render(request, 'reporting/risk_selection.html', context)


@staff_member_required
def selection_summary_report(request, engine):
    """Selection Summary Report"""
    from_period = request.GET.get('from_tax_period')
    to_period = request.GET.get('to_tax_period')
    
    queryset = ComplianceRiskReferral.objects.all()
    if from_period:
        queryset = queryset.filter(assessment_from_period__gte=from_period)
    if to_period:
        queryset = queryset.filter(assessment_to_period__lte=to_period)
    
    summary = {
        'audit': queryset.filter(final_selection='AUDIT').count(),
        'review': queryset.filter(final_selection='REVIEW').count(),
        'monitor': queryset.filter(final_selection='MONITOR').count(),
        'not_selected': queryset.filter(final_selection='NOT SELECTED').count(),
    }
    
    context = {
        'report_title': 'Selection Summary',
        'summary': summary,
        'report_period': f"{from_period or 'All'} to {to_period or 'All'}",
        'filters': {
            'from_tax_period': from_period,
            'to_tax_period': to_period,
            'tax_periods': engine.tax_periods,
        },
        'generated_date': timezone.now().strftime('%d-%m-%Y %H:%M'),
        'generated_by': request.user.username,
    }
    
    return render(request, 'reporting/selection_summary.html', context)

@staff_member_required
def selection_summary_report(request, engine):
    """Selection Summary Report"""
    from_period = request.GET.get('from_tax_period')
    to_period = request.GET.get('to_tax_period')
    
    queryset = ComplianceRiskReferral.objects.all()
    if from_period:
        queryset = queryset.filter(assessment_from_period__gte=from_period)
    if to_period:
        queryset = queryset.filter(assessment_to_period__lte=to_period)
    
    summary = {
        'audit': queryset.filter(final_selection='AUDIT').count(),
        'review': queryset.filter(final_selection='REVIEW').count(),
        'monitor': queryset.filter(final_selection='MONITOR').count(),
        'not_selected': queryset.filter(final_selection='NOT SELECTED').count(),
    }
    
    context = {
        'report_title': 'Selection Summary',
        'summary': summary,
        'report_period': f"{from_period or 'All'} to {to_period or 'All'}",
        'filters': {
            'from_tax_period': from_period,
            'to_tax_period': to_period,
            'tax_periods': engine.tax_periods,
        },
        'generated_date': timezone.now().strftime('%d-%m-%Y %H:%M'),
        'generated_by': request.user.username,
    }
    
    return render(request, 'reporting/selection_summary.html', context)


# Audit Reports Implementation
@staff_member_required
def audit_case_register_report(request, engine):
    """Audit Case Register Report"""
    from_period = request.GET.get('from_tax_period')
    to_period = request.GET.get('to_tax_period')
    status = request.GET.get('status')
    gstin = request.GET.get('gstin')
    
    queryset = AuditCase.objects.all()
    if from_period:
        queryset = queryset.filter(from_tax_period__gte=from_period)
    if to_period:
        queryset = queryset.filter(to_tax_period__lte=to_period)
    if status:
        queryset = queryset.filter(status=status)
    if gstin:
        queryset = queryset.filter(gstin__icontains=gstin)
    
    audit_cases = []
    for case in queryset:
        audit_cases.append({
            'audit_case_id': case.audit_case_id,
            'gstin': case.gstin,
            'taxpayer_name': case.taxpayer_name,
            'from_tax_period': case.from_tax_period,
            'to_tax_period': case.to_tax_period,
            'assessment_type': case.assessment_type,
            'status': case.status,
            'assigned_officer': case.assigned_officer.username if case.assigned_officer else '-',
            'created_date': case.created_at,
            'id': case.id,
        })
    
    # Pagination
    page = request.GET.get('page', 1)
    paginator = Paginator(audit_cases, 50)
    audit_list = paginator.get_page(page)
    
    context = {
        'report_title': 'Audit Case Register',
        'audit_cases': audit_list,
        'filters': {
            'from_tax_period': from_period,
            'to_tax_period': to_period,
            'status': status,
            'gstin': gstin,
            'tax_periods': engine.tax_periods,
        },
        'filter_options': {
            'statuses': ['Pending Assignment', 'Assigned', 'In Progress', 'Completed', 'Closed'],
        },
        'generated_date': timezone.now().strftime('%d-%m-%Y %H:%M'),
        'generated_by': request.user.username,
    }
    
    return render(request, 'reporting/audit_case_register.html', context)


@staff_member_required
def audit_status_report(request, engine):
    """Audit Status Summary Report"""
    from_period = request.GET.get('from_tax_period')
    to_period = request.GET.get('to_tax_period')
    
    queryset = AuditCase.objects.all()
    if from_period:
        queryset = queryset.filter(from_tax_period__gte=from_period)
    if to_period:
        queryset = queryset.filter(to_tax_period__lte=to_period)
    
    summary = {
        'pending': queryset.filter(status='Pending Assignment').count(),
        'assigned': queryset.filter(status='Assigned').count(),
        'in_progress': queryset.filter(status='In Progress').count(),
        'completed': queryset.filter(status='Completed').count(),
        'closed': queryset.filter(status='Closed').count(),
    }
    
    context = {
        'report_title': 'Audit Status Summary',
        'summary': summary,
        'report_period': f"{from_period or 'All'} to {to_period or 'All'}",
        'filters': {
            'from_tax_period': from_period,
            'to_tax_period': to_period,
            'tax_periods': engine.tax_periods,
        },
        'generated_date': timezone.now().strftime('%d-%m-%Y %H:%M'),
        'generated_by': request.user.username,
    }
    
    return render(request, 'reporting/audit_status.html', context)


@staff_member_required
def audit_workload_report(request, engine):
    """Audit Officer Workload Report"""
    queryset = AuditCase.objects.all()
    
    # Calculate workload per officer
    officer_workload = []
    officers = User.objects.filter(groups__name='Audit Officers')
    
    for officer in officers:
        cases = queryset.filter(assigned_officer=officer)
        officer_workload.append({
            'officer': officer.username,
            'total_cases': cases.count(),
            'pending': cases.filter(status='Pending Assignment').count(),
            'in_progress': cases.filter(status='In Progress').count(),
            'completed': cases.filter(status='Completed').count(),
            'closed': cases.filter(status='Closed').count(),
        })
    
    context = {
        'report_title': 'Audit Officer Workload',
        'officer_workload': officer_workload,
        'generated_date': timezone.now().strftime('%d-%m-%Y %H:%M'),
        'generated_by': request.user.username,
    }
    
    return render(request, 'reporting/audit_workload.html', context)


@staff_member_required
def audit_findings_report(request, engine):
    """Audit Findings Report"""
    from_period = request.GET.get('from_tax_period')
    to_period = request.GET.get('to_tax_period')
    
    queryset = AuditFinding.objects.all()
    if from_period:
        queryset = queryset.filter(assessment__from_tax_period__gte=from_period)
    if to_period:
        queryset = queryset.filter(assessment__to_tax_period__lte=to_period)
    
    findings = []
    for finding in queryset:
        findings.append({
            'audit_case_id': finding.assessment.audit_case.audit_case_id if finding.assessment.audit_case else '-',
            'gstin': finding.assessment.gstin if finding.assessment else '-',
            'taxpayer_name': finding.assessment.taxpayer_name if finding.assessment else '-',
            'finding_type': finding.finding_type,
            'description': finding.description,
            'shortfall_amount': engine.format_number(finding.shortfall_amount or 0),
            'corrective_action': finding.corrective_action,
            'status': finding.status,
            'id': finding.id,
        })
    
    # Pagination
    page = request.GET.get('page', 1)
    paginator = Paginator(findings, 50)
    findings_list = paginator.get_page(page)
    
    context = {
        'report_title': 'Audit Findings Report',
        'findings': findings_list,
        'report_period': f"{from_period or 'All'} to {to_period or 'All'}",
        'filters': {
            'from_tax_period': from_period,
            'to_tax_period': to_period,
            'tax_periods': engine.tax_periods,
        },
        'generated_date': timezone.now().strftime('%d-%m-%Y %H:%M'),
        'generated_by': request.user.username,
    }
    
    return render(request, 'reporting/audit_findings.html', context)


@staff_member_required
def audit_outcome_report(request, engine):
    """Audit Outcome Report"""
    from_period = request.GET.get('from_tax_period')
    to_period = request.GET.get('to_tax_period')
    
    queryset = AuditAssessment.objects.all()
    if from_period:
        queryset = queryset.filter(from_tax_period__gte=from_period)
    if to_period:
        queryset = queryset.filter(to_tax_period__lte=to_period)
    
    outcomes = []
    for assessment in queryset:
        outcomes.append({
            'audit_case_id': assessment.audit_case.audit_case_id if assessment.audit_case else '-',
            'gstin': assessment.gstin,
            'taxpayer_name': assessment.taxpayer_name,
            'tax_period': assessment.tax_period,
            'assessment_type': assessment.assessment_type,
            'outcome': assessment.outcome,
            'assessment_date': assessment.assessment_date,
            'id': assessment.id,
        })
    
    # Pagination
    page = request.GET.get('page', 1)
    paginator = Paginator(outcomes, 50)
    outcomes_list = paginator.get_page(page)
    
    context = {
        'report_title': 'Audit Outcome Report',
        'outcomes': outcomes_list,
        'report_period': f"{from_period or 'All'} to {to_period or 'All'}",
        'filters': {
            'from_tax_period': from_period,
            'to_tax_period': to_period,
            'tax_periods': engine.tax_periods,
        },
        'generated_date': timezone.now().strftime('%d-%m-%Y %H:%M'),
        'generated_by': request.user.username,
    }
    
    return render(request, 'reporting/audit_outcome.html', context)


@staff_member_required
def audit_assessment_report(request, engine):
    """Audit Assessment Report"""
    from_period = request.GET.get('from_tax_period')
    to_period = request.GET.get('to_tax_period')
    
    queryset = AuditAssessment.objects.all()
    if from_period:
        queryset = queryset.filter(from_tax_period__gte=from_period)
    if to_period:
        queryset = queryset.filter(to_tax_period__lte=to_period)
    
    assessments = []
    for assessment in queryset:
        assessments.append({
            'audit_case_id': assessment.audit_case.audit_case_id if assessment.audit_case else '-',
            'gstin': assessment.gstin,
            'taxpayer_name': assessment.taxpayer_name,
            'tax_period': assessment.tax_period,
            'declared_sales': engine.format_number(assessment.declared_sales or 0),
            'assessed_sales': engine.format_number(assessment.assessed_sales or 0),
            'declared_output_gst': engine.format_number(assessment.declared_output_gst or 0),
            'assessed_output_gst': engine.format_number(assessment.assessed_output_gst or 0),
            'difference': engine.format_number((assessment.assessed_output_gst or 0) - (assessment.declared_output_gst or 0)),
            'id': assessment.id,
        })
    
    # Pagination
    page = request.GET.get('page', 1)
    paginator = Paginator(assessments, 50)
    assessments_list = paginator.get_page(page)
    
    context = {
        'report_title': 'Audit Assessment Report',
        'assessments': assessments_list,
        'report_period': f"{from_period or 'All'} to {to_period or 'All'}",
        'filters': {
            'from_tax_period': from_period,
            'to_tax_period': to_period,
            'tax_periods': engine.tax_periods,
        },
        'generated_date': timezone.now().strftime('%d-%m-%Y %H:%M'),
        'generated_by': request.user.username,
    }
    
    return render(request, 'reporting/audit_assessment.html', context)


@staff_member_required
def audit_revenue_report(request, engine):
    """Audit Revenue Impact Report"""
    from_period = request.GET.get('from_tax_period')
    to_period = request.GET.get('to_tax_period')
    
    queryset = AuditAssessment.objects.all()
    if from_period:
        queryset = queryset.filter(from_tax_period__gte=from_period)
    if to_period:
        queryset = queryset.filter(to_tax_period__lte=to_period)
    
    # Calculate revenue impact
    total_assessed = queryset.aggregate(Sum('assessed_output_gst'))['assessed_output_gst__sum'] or 0
    total_declared = queryset.aggregate(Sum('declared_output_gst'))['declared_output_gst__sum'] or 0
    total_shortfall = queryset.aggregate(Sum('shortfall_amount'))['shortfall_amount__sum'] or 0
    total_recovered = queryset.aggregate(Sum('amount_recovered'))['amount_recovered__sum'] or 0
    
    context = {
        'report_title': 'Audit Revenue Impact',
        'revenue_impact': {
            'total_assessed': engine.format_number(total_assessed),
            'total_declared': engine.format_number(total_declared),
            'total_shortfall': engine.format_number(total_shortfall),
            'total_recovered': engine.format_number(total_recovered),
            'recovery_rate': engine.calculate_percentage(total_recovered, total_shortfall),
        },
        'report_period': f"{from_period or 'All'} to {to_period or 'All'}",
        'filters': {
            'from_tax_period': from_period,
            'to_tax_period': to_period,
            'tax_periods': engine.tax_periods,
        },
        'generated_date': timezone.now().strftime('%d-%m-%Y %H:%M'),
        'generated_by': request.user.username,
    }
    
    return render(request, 'reporting/audit_revenue.html', context)


# Refund Reports Implementation
@staff_member_required
def refund_register_report(request, engine):
    """Refund Register Report"""
    from_period = request.GET.get('from_tax_period')
    to_period = request.GET.get('to_tax_period')
    status = request.GET.get('status')
    gstin = request.GET.get('gstin')
    
    queryset = RefundRegister.objects.all()
    if from_period:
        queryset = queryset.filter(tax_period__gte=from_period)
    if to_period:
        queryset = queryset.filter(tax_period__lte=to_period)
    if status:
        queryset = queryset.filter(status=status)
    if gstin:
        queryset = queryset.filter(gst_tpn__icontains=gstin)
    
    refunds = []
    for refund in queryset:
        refunds.append({
            'refund_id': refund.refund_id,
            'gst_tpn': refund.gst_tpn,
            'taxpayer_name': refund.taxpayer_name,
            'tax_period': refund.tax_period,
            'claimed_amount': engine.format_number(refund.claimed_amount or 0),
            'refund_approved': engine.format_number(refund.refund_approved or 0),
            'adjustment': engine.format_number(refund.adjustment or 0),
            'status': refund.status,
            'submitted_date': refund.submitted_date,
            'id': refund.id,
        })
    
    # Pagination
    page = request.GET.get('page', 1)
    paginator = Paginator(refunds, 50)
    refund_list = paginator.get_page(page)
    
    context = {
        'report_title': 'Refund Register',
        'refunds': refund_list,
        'filters': {
            'from_tax_period': from_period,
            'to_tax_period': to_period,
            'status': status,
            'gstin': gstin,
            'tax_periods': engine.tax_periods,
        },
        'filter_options': {
            'statuses': ['submitted', 'under_review', 'processing', 'approved', 'rejected', 'paid', 'closed'],
        },
        'generated_date': timezone.now().strftime('%d-%m-%Y %H:%M'),
        'generated_by': request.user.username,
    }
    
    return render(request, 'reporting/refund_register.html', context)


@staff_member_required
def refund_status_report(request, engine):
    """Refund Status Summary Report"""
    from_period = request.GET.get('from_tax_period')
    to_period = request.GET.get('to_tax_period')
    
    queryset = RefundRegister.objects.all()
    if from_period:
        queryset = queryset.filter(tax_period__gte=from_period)
    if to_period:
        queryset = queryset.filter(tax_period__lte=to_period)
    
    summary = {
        'submitted': queryset.filter(status='submitted').count(),
        'under_review': queryset.filter(status='under_review').count(),
        'processing': queryset.filter(status='processing').count(),
        'approved': queryset.filter(status='approved').count(),
        'rejected': queryset.filter(status='rejected').count(),
        'paid': queryset.filter(status='paid').count(),
        'closed': queryset.filter(status='closed').count(),
    }
    
    context = {
        'report_title': 'Refund Status Summary',
        'summary': summary,
        'report_period': f"{from_period or 'All'} to {to_period or 'All'}",
        'filters': {
            'from_tax_period': from_period,
            'to_tax_period': to_period,
            'tax_periods': engine.tax_periods,
        },
        'generated_date': timezone.now().strftime('%d-%m-%Y %H:%M'),
        'generated_by': request.user.username,
    }
    
    return render(request, 'reporting/refund_status.html', context)


@staff_member_required
def refund_amount_report(request, engine):
    """Refund Amount Summary Report"""
    from_period = request.GET.get('from_tax_period')
    to_period = request.GET.get('to_tax_period')
    
    queryset = RefundRegister.objects.all()
    if from_period:
        queryset = queryset.filter(tax_period__gte=from_period)
    if to_period:
        queryset = queryset.filter(tax_period__lte=to_period)
    
    summary = queryset.aggregate(
        total_claimed=Sum('claimed_amount'),
        total_approved=Sum('refund_approved'),
        total_adjusted=Sum('adjustment'),
        total_paid=Sum('refund_approved')  # Using approved as proxy for paid
    )
    
    context = {
        'report_title': 'Refund Amount Summary',
        'summary': {
            'total_claimed': engine.format_number(summary['total_claimed'] or 0),
            'total_approved': engine.format_number(summary['total_approved'] or 0),
            'total_adjusted': engine.format_number(summary['total_adjusted'] or 0),
            'total_paid': engine.format_number(summary['total_paid'] or 0),
        },
        'report_period': f"{from_period or 'All'} to {to_period or 'All'}",
        'filters': {
            'from_tax_period': from_period,
            'to_tax_period': to_period,
            'tax_periods': engine.tax_periods,
        },
        'generated_date': timezone.now().strftime('%d-%m-%Y %H:%M'),
        'generated_by': request.user.username,
    }
    
    return render(request, 'reporting/refund_amount.html', context)


@staff_member_required
def refund_processing_report(request, engine):
    """Refund Processing Time Report"""
    from_period = request.GET.get('from_tax_period')
    to_period = request.GET.get('to_tax_period')
    
    queryset = RefundRegister.objects.filter(status__in=['approved', 'paid', 'closed'])
    if from_period:
        queryset = queryset.filter(tax_period__gte=from_period)
    if to_period:
        queryset = queryset.filter(tax_period__lte=to_period)
    
    processing_times = []
    for refund in queryset:
        processing_days = None
        if refund.submitted_date and refund.approved_date:
            processing_days = (refund.approved_date - refund.submitted_date).days
        
        processing_times.append({
            'refund_id': refund.refund_id,
            'gst_tpn': refund.gst_tpn,
            'taxpayer_name': refund.taxpayer_name,
            'tax_period': refund.tax_period,
            'submitted_date': refund.submitted_date,
            'approved_date': refund.approved_date,
            'processing_days': processing_days,
            'status': refund.status,
            'id': refund.id,
        })
    
    # Pagination
    page = request.GET.get('page', 1)
    paginator = Paginator(processing_times, 50)
    processing_list = paginator.get_page(page)
    
    context = {
        'report_title': 'Refund Processing Time',
        'processing_times': processing_list,
        'report_period': f"{from_period or 'All'} to {to_period or 'All'}",
        'filters': {
            'from_tax_period': from_period,
            'to_tax_period': to_period,
            'tax_periods': engine.tax_periods,
        },
        'generated_date': timezone.now().strftime('%d-%m-%Y %H:%M'),
        'generated_by': request.user.username,
    }
    
    return render(request, 'reporting/refund_processing.html', context)


@staff_member_required
def refund_by_period_report(request, engine):
    """Refund by Tax Period Report"""
    from_period = request.GET.get('from_tax_period')
    to_period = request.GET.get('to_tax_period')
    
    queryset = RefundRegister.objects.all()
    if from_period:
        queryset = queryset.filter(tax_period__gte=from_period)
    if to_period:
        queryset = queryset.filter(tax_period__lte=to_period)
    
    period_summary = queryset.values('tax_period').annotate(
        count=Count('id'),
        total_claimed=Sum('claimed_amount'),
        total_approved=Sum('refund_approved')
    ).order_by('tax_period')
    
    for item in period_summary:
        item['total_claimed'] = engine.format_number(item['total_claimed'] or 0)
        item['total_approved'] = engine.format_number(item['total_approved'] or 0)
    
    context = {
        'report_title': 'Refund by Tax Period',
        'period_summary': period_summary,
        'report_period': f"{from_period or 'All'} to {to_period or 'All'}",
        'filters': {
            'from_tax_period': from_period,
            'to_tax_period': to_period,
            'tax_periods': engine.tax_periods,
        },
        'generated_date': timezone.now().strftime('%d-%m-%Y %H:%M'),
        'generated_by': request.user.username,
    }
    
    return render(request, 'reporting/refund_by_period.html', context)


# Enforcement Reports Implementation
@staff_member_required
def enforcement_register_report(request, engine):
    """Enforcement Case Register Report"""
    from_period = request.GET.get('from_tax_period')
    to_period = request.GET.get('to_tax_period')
    status = request.GET.get('status')
    gstin = request.GET.get('gstin')
    
    queryset = EnforcementRecovery.objects.all()
    if from_period:
        queryset = queryset.filter(tax_period__gte=from_period)
    if to_period:
        queryset = queryset.filter(tax_period__lte=to_period)
    if status:
        queryset = queryset.filter(status=status)
    if gstin:
        queryset = queryset.filter(gstin__icontains=gstin)
    
    enforcement_cases = []
    for case in queryset:
        outstanding = (case.amount_due or 0) - (case.amount_recovered or 0)
        
        enforcement_cases.append({
            'case_id': case.case_id,
            'gstin': case.gstin,
            'taxpayer_name': case.taxpayer_name,
            'tax_period': case.tax_period,
            'case_type': case.case_type,
            'amount_due': engine.format_number(case.amount_due or 0),
            'amount_recovered': engine.format_number(case.amount_recovered or 0),
            'outstanding': engine.format_number(outstanding),
            'status': case.status,
            'assigned_officer': case.assigned_officer.username if case.assigned_officer else '-',
            'notice_date': case.notice_date,
            'id': case.id,
        })
    
    # Pagination
    page = request.GET.get('page', 1)
    paginator = Paginator(enforcement_cases, 50)
    enforcement_list = paginator.get_page(page)
    
    context = {
        'report_title': 'Enforcement Case Register',
        'enforcement_cases': enforcement_list,
        'filters': {
            'from_tax_period': from_period,
            'to_tax_period': to_period,
            'status': status,
            'gstin': gstin,
            'tax_periods': engine.tax_periods,
        },
        'filter_options': {
            'statuses': ['Initiated', 'Notice Issued', 'Assessment Issued', 'Recovery Initiated', 'Recovered', 'Closed'],
        },
        'generated_date': timezone.now().strftime('%d-%m-%Y %H:%M'),
        'generated_by': request.user.username,
    }
    
    return render(request, 'reporting/enforcement_register.html', context)


@staff_member_required
def outstanding_recovery_report(request, engine):
    """Outstanding Recovery Report"""
    from_period = request.GET.get('from_tax_period')
    to_period = request.GET.get('to_tax_period')
    
    queryset = EnforcementRecovery.objects.exclude(status__in=['Recovered', 'Closed'])
    if from_period:
        queryset = queryset.filter(tax_period__gte=from_period)
    if to_period:
        queryset = queryset.filter(tax_period__lte=to_period)
    
    outstanding_cases = []
    for case in queryset:
        outstanding = (case.amount_due or 0) - (case.amount_recovered or 0)
        
        outstanding_cases.append({
            'case_id': case.case_id,
            'gstin': case.gstin,
            'taxpayer_name': case.taxpayer_name,
            'tax_period': case.tax_period,
            'amount_due': engine.format_number(case.amount_due or 0),
            'amount_recovered': engine.format_number(case.amount_recovered or 0),
            'outstanding': engine.format_number(outstanding),
            'status': case.status,
            'notice_date': case.notice_date,
            'assigned_officer': case.assigned_officer.username if case.assigned_officer else '-',
            'id': case.id,
        })
    
    # Pagination
    page = request.GET.get('page', 1)
    paginator = Paginator(outstanding_cases, 50)
    outstanding_list = paginator.get_page(page)
    
    context = {
        'report_title': 'Outstanding Recovery Report',
        'outstanding_cases': outstanding_list,
        'report_period': f"{from_period or 'All'} to {to_period or 'All'}",
        'filters': {
            'from_tax_period': from_period,
            'to_tax_period': to_period,
            'tax_periods': engine.tax_periods,
        },
        'generated_date': timezone.now().strftime('%d-%m-%Y %H:%M'),
        'generated_by': request.user.username,
    }
    
    return render(request, 'reporting/outstanding_recovery.html', context)


@staff_member_required
def recovery_summary_report(request, engine):
    """Recovery Summary Report"""
    from_period = request.GET.get('from_tax_period')
    to_period = request.GET.get('to_tax_period')
    
    queryset = EnforcementRecovery.objects.all()
    if from_period:
        queryset = queryset.filter(tax_period__gte=from_period)
    if to_period:
        queryset = queryset.filter(tax_period__lte=to_period)
    
    summary = queryset.aggregate(
        total_cases=Count('id'),
        total_amount_due=Sum('amount_due'),
        total_amount_recovered=Sum('amount_recovered')
    )
    
    total_outstanding = (summary['total_amount_due'] or 0) - (summary['total_amount_recovered'] or 0)
    recovery_rate = engine.calculate_percentage(summary['total_amount_recovered'], summary['total_amount_due'])
    
    context = {
        'report_title': 'Recovery Summary',
        'summary': {
            'total_cases': summary['total_cases'],
            'total_amount_due': engine.format_number(summary['total_amount_due'] or 0),
            'total_amount_recovered': engine.format_number(summary['total_amount_recovered'] or 0),
            'total_outstanding': engine.format_number(total_outstanding),
            'recovery_rate': recovery_rate,
        },
        'report_period': f"{from_period or 'All'} to {to_period or 'All'}",
        'filters': {
            'from_tax_period': from_period,
            'to_tax_period': to_period,
            'tax_periods': engine.tax_periods,
        },
        'generated_date': timezone.now().strftime('%d-%m-%Y %H:%M'),
        'generated_by': request.user.username,
    }
    
    return render(request, 'reporting/recovery_summary.html', context)


@staff_member_required
def non_payment_report(request, engine):
    """Non-Payment Case Report"""
    from_period = request.GET.get('from_tax_period')
    to_period = request.GET.get('to_tax_period')
    
    # Get returns with payment defaults
    queryset = GSTReturn.objects.filter(payment_status='Default')
    if from_period:
        queryset = queryset.filter(tax_period__gte=from_period)
    if to_period:
        queryset = queryset.filter(tax_period__lte=to_period)
    
    non_payment_cases = []
    for return_obj in queryset:
        gst_payable = return_obj.gst_payable_refundable or 0
        actual_payment = return_obj.actual_gst_payment_received or 0
        outstanding = gst_payable - actual_payment if gst_payable > 0 else 0
        
        non_payment_cases.append({
            'tax_period': return_obj.tax_period,
            'gstin': return_obj.gstin,
            'taxpayer_name': return_obj.taxpayer_name,
            'gst_payable': engine.format_number(gst_payable),
            'actual_payment': engine.format_number(actual_payment),
            'outstanding': engine.format_number(outstanding),
            'payment_status': return_obj.payment_status,
            'compliance_status': return_obj.compliance_status,
            'id': return_obj.id,
        })
    
    # Pagination
    page = request.GET.get('page', 1)
    paginator = Paginator(non_payment_cases, 50)
    non_payment_list = paginator.get_page(page)
    
    context = {
        'report_title': 'Non-Payment Case Report',
        'non_payment_cases': non_payment_list,
        'report_period': f"{from_period or 'All'} to {to_period or 'All'}",
        'filters': {
            'from_tax_period': from_period,
            'to_tax_period': to_period,
            'tax_periods': engine.tax_periods,
        },
        'generated_date': timezone.now().strftime('%d-%m-%Y %H:%M'),
        'generated_by': request.user.username,
    }
    
    return render(request, 'reporting/non_payment.html', context)


@staff_member_required
def enforcement_status_report(request, engine):
    """Enforcement Status Report"""
    from_period = request.GET.get('from_tax_period')
    to_period = request.GET.get('to_tax_period')
    
    queryset = EnforcementRecovery.objects.all()
    if from_period:
        queryset = queryset.filter(tax_period__gte=from_period)
    if to_period:
        queryset = queryset.filter(tax_period__lte=to_period)
    
    summary = {
        'initiated': queryset.filter(status='Initiated').count(),
        'notice_issued': queryset.filter(status='Notice Issued').count(),
        'assessment_issued': queryset.filter(status='Assessment Issued').count(),
        'recovery_initiated': queryset.filter(status='Recovery Initiated').count(),
        'recovered': queryset.filter(status='Recovered').count(),
        'closed': queryset.filter(status='Closed').count(),
    }
    
    context = {
        'report_title': 'Enforcement Status Report',
        'summary': summary,
        'report_period': f"{from_period or 'All'} to {to_period or 'All'}",
        'filters': {
            'from_tax_period': from_period,
            'to_tax_period': to_period,
            'tax_periods': engine.tax_periods,
        },
        'generated_date': timezone.now().strftime('%d-%m-%Y %H:%M'),
        'generated_by': request.user.username,
    }
    
    return render(request, 'reporting/enforcement_status.html', context)


# Officer/Workload Reports Implementation
@staff_member_required
def officer_workload_report(request, engine):
    """Officer Workload Report"""
    officer_workload = []
    
    # Get audit officers
    audit_officers = User.objects.filter(groups__name='Audit Officers')
    for officer in audit_officers:
        audit_cases = AuditCase.objects.filter(assigned_officer=officer)
        officer_workload.append({
            'officer': officer.username,
            'type': 'Audit',
            'total_cases': audit_cases.count(),
            'pending': audit_cases.filter(status='Pending Assignment').count(),
            'in_progress': audit_cases.filter(status='In Progress').count(),
            'completed': audit_cases.filter(status='Completed').count(),
        })
    
    # Get enforcement officers
    enforcement_officers = User.objects.filter(groups__name='Enforcement Officers')
    for officer in enforcement_officers:
        enforcement_cases = EnforcementRecovery.objects.filter(assigned_officer=officer)
        officer_workload.append({
            'officer': officer.username,
            'type': 'Enforcement',
            'total_cases': enforcement_cases.count(),
            'pending': enforcement_cases.filter(status='Initiated').count(),
            'in_progress': enforcement_cases.filter(status__in=['Notice Issued', 'Assessment Issued', 'Recovery Initiated']).count(),
            'completed': enforcement_cases.filter(status__in=['Recovered', 'Closed']).count(),
        })
    
    context = {
        'report_title': 'Officer Workload Report',
        'officer_workload': officer_workload,
        'generated_date': timezone.now().strftime('%d-%m-%Y %H:%M'),
        'generated_by': request.user.username,
    }
    
    return render(request, 'reporting/officer_workload.html', context)


@staff_member_required
def audit_performance_report(request, engine):
    """Audit Performance Report"""
    from_period = request.GET.get('from_tax_period')
    to_period = request.GET.get('to_tax_period')
    
    queryset = AuditAssessment.objects.all()
    if from_period:
        queryset = queryset.filter(from_tax_period__gte=from_period)
    if to_period:
        queryset = queryset.filter(to_tax_period__lte=to_period)
    
    performance_data = []
    officers = User.objects.filter(groups__name='Audit Officers')
    
    for officer in officers:
        assessments = queryset.filter(assessor=officer)
        total = assessments.count()
        completed = assessments.filter(outcome__in=['Assessment Completed', 'Shortfall Confirmed']).count()
        
        performance_data.append({
            'officer': officer.username,
            'total_assessments': total,
            'completed': completed,
            'completion_rate': engine.calculate_percentage(completed, total),
        })
    
    context = {
        'report_title': 'Audit Performance Report',
        'performance_data': performance_data,
        'report_period': f"{from_period or 'All'} to {to_period or 'All'}",
        'filters': {
            'from_tax_period': from_period,
            'to_tax_period': to_period,
            'tax_periods': engine.tax_periods,
        },
        'generated_date': timezone.now().strftime('%d-%m-%Y %H:%M'),
        'generated_by': request.user.username,
    }
    
    return render(request, 'reporting/audit_performance.html', context)


@staff_member_required
def refund_performance_report(request, engine):
    """Refund Processing Performance Report"""
    from_period = request.GET.get('from_tax_period')
    to_period = request.GET.get('to_tax_period')
    
    queryset = RefundRegister.objects.filter(status__in=['approved', 'paid', 'closed'])
    if from_period:
        queryset = queryset.filter(tax_period__gte=from_period)
    if to_period:
        queryset = queryset.filter(tax_period__lte=to_period)
    
    total = queryset.count()
    avg_processing_days = 0
    
    processing_days_list = []
    for refund in queryset:
        if refund.submitted_date and refund.approved_date:
            days = (refund.approved_date - refund.submitted_date).days
            processing_days_list.append(days)
    
    if processing_days_list:
        avg_processing_days = sum(processing_days_list) / len(processing_days_list)
    
    context = {
        'report_title': 'Refund Processing Performance',
        'performance': {
            'total_processed': total,
            'avg_processing_days': round(avg_processing_days, 2),
        },
        'report_period': f"{from_period or 'All'} to {to_period or 'All'}",
        'filters': {
            'from_tax_period': from_period,
            'to_tax_period': to_period,
            'tax_periods': engine.tax_periods,
        },
        'generated_date': timezone.now().strftime('%d-%m-%Y %H:%M'),
        'generated_by': request.user.username,
    }
    
    return render(request, 'reporting/refund_performance.html', context)


# Custom Report Builder
@staff_member_required
def custom_report_builder(request, engine):
    """Custom Report Builder"""
    # Get all available models and their fields
    data_sources = {
        'taxpayers': {
            'model': 'TaxpayerMaster',
            'fields': ['gstin', 'taxpayer_name', 'business_name', 'organisation_type', 'sector', 'dzongkhag', 'frequency', 'status']
        },
        'returns': {
            'model': 'GSTReturn',
            'fields': ['gstin', 'taxpayer_name', 'tax_period', 'declared_sales', 'total_itc_claimed', 'gst_payable_refundable', 'filing_status', 'payment_status']
        },
        'compliance': {
            'model': 'ComplianceMonitoring',
            'fields': ['gstin', 'taxpayer_name', 'tax_period', 'compliance_status', 'monitoring_notes']
        },
        'risk': {
            'model': 'ComplianceRiskReferral',
            'fields': ['risk_id', 'gstin', 'taxpayer_name', 'assessment_from_period', 'assessment_to_period', 'risk_level', 'risk_type', 'final_selection']
        },
        'audit': {
            'model': 'AuditCase',
            'fields': ['audit_case_id', 'gstin', 'taxpayer_name', 'from_tax_period', 'to_tax_period', 'assessment_type', 'status', 'assigned_officer']
        },
        'refund': {
            'model': 'RefundRegister',
            'fields': ['refund_id', 'gst_tpn', 'taxpayer_name', 'tax_period', 'claimed_amount', 'refund_approved', 'status']
        },
    }
    
    context = {
        'report_title': 'Custom Report Builder',
        'data_sources': data_sources,
        'generated_date': timezone.now().strftime('%d-%m-%Y %H:%M'),
        'generated_by': request.user.username,
    }
    
    return render(request, 'reporting/custom_builder.html', context)

@staff_member_required
def refund_performance_report(request, engine):
    """Refund Processing Performance Report"""
    from_period = request.GET.get('from_tax_period')
    to_period = request.GET.get('to_tax_period')
    
    queryset = RefundRegister.objects.filter(status__in=['approved', 'paid', 'closed'])
    if from_period:
        queryset = queryset.filter(tax_period__gte=from_period)
    if to_period:
        queryset = queryset.filter(tax_period__lte=to_period)
    
    total = queryset.count()
    avg_processing_days = 0
    
    processing_days_list = []
    for refund in queryset:
        if refund.submitted_date and refund.approved_date:
            days = (refund.approved_date - refund.submitted_date).days
            processing_days_list.append(days)
    
    if processing_days_list:
        avg_processing_days = sum(processing_days_list) / len(processing_days_list)
    
    context = {
        'report_title': 'Refund Processing Performance',
        'performance': {
            'total_processed': total,
            'avg_processing_days': round(avg_processing_days, 2),
        },
        'report_period': f"{from_period or 'All'} to {to_period or 'All'}",
        'filters': {
            'from_tax_period': from_period,
            'to_tax_period': to_period,
            'tax_periods': engine.tax_periods,
        },
        'generated_date': timezone.now().strftime('%d-%m-%Y %H:%M'),
        'generated_by': request.user.username,
    }
    
    return render(request, 'reporting/refund_performance.html', context)


# Export Functions
@staff_member_required
def export_excel(request, report_id):
    """Export report to Excel"""
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{report_id}_report.xlsx"'
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"
    
    # Add headers
    ws.append(['Report Name', report_id])
    ws.append(['Generated Date', timezone.now().strftime('%d-%m-%Y %H:%M')])
    ws.append(['Generated By', request.user.username])
    ws.append([])
    
    # Add basic styling
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    
    # Style the header row
    for cell in ws[1:1]:
        cell.font = header_font
        cell.fill = header_fill
    
    # Save to response
    wb.save(response)
    return response


@staff_member_required
def export_pdf(request, report_id):
    """Export report to PDF"""
    # For now, return a placeholder response
    # In production, this would use reportlab or similar
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{report_id}_report.pdf"'
    
    # Simple PDF placeholder
    response.write(f"PDF export for {report_id} - To be implemented with reportlab".encode())
    return response


@staff_member_required
def export_csv(request, report_id):
    """Export report to CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{report_id}_report.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Report Name', report_id])
    writer.writerow(['Generated Date', timezone.now().strftime('%d-%m-%Y %H:%M')])
    writer.writerow(['Generated By', request.user.username])
    writer.writerow([])
    
    return response
